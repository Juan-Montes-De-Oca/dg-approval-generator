"""
DG APPROVAL GENERATOR — Web App
Hecho con Streamlit · Proyecto de Juan Deo
"""

import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Side, Border
from openpyxl.utils import get_column_letter
from datetime import datetime
import io, os

# ── Page config ───────────────────────────────────────────────────────────
st.set_page_config(
    page_title="DG Approval Generator",
    page_icon="🚢",
    layout="wide"
)

# ── Styles ─────────────────────────────────────────────────────────────
st.markdown("""
<style>
    .main { background-color: #f7fbfe; }
    .stButton>button {
        background-color: #1E6091;
        color: white;
        font-weight: bold;
        border-radius: 8px;
        border: none;
        padding: 0.6rem 2rem;
        font-size: 16px;
    }
    .stButton>button:hover { background-color: #2E5090; }
    .success-box {
        background: #d4edda; border-left: 5px solid #27AE60;
        padding: 1rem; border-radius: 8px; margin: 1rem 0;
    }
    .warning-box {
        background: #fff3cd; border-left: 5px solid #F0A500;
        padding: 1rem; border-radius: 8px; margin: 1rem 0;
    }
</style>
""", unsafe_allow_html=True)

# ── Header ─────────────────────────────────────────────────────────────
st.markdown("## 🚢 DG Approval Generator")
st.markdown("Sube tu DCR, configura el barco y descarga las plantillas listas.")
st.divider()

# ── Sidebar config ──────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### ⚙️ Configuración")
    vessel_name = st.text_input(
        "Nombre del barco",
        value="AS SVENJA / V.26022N",
        help="Nombre y voyage tal como aparece en el punto 2"
    )
    port_loading = st.text_input(
        "Port of Loading",
        value="BALBOA",
        help="Puerto de carga (punto 5)"
    )
    st.divider()
    st.markdown("**Reglas aplicadas:**")
    st.markdown("""
    - `*` → **N** (no marine pollutant)
    - `P` → **P** (marine pollutant)
    - Gross Weight vacío → estimado (NW × 1.05)
    - Port of Loading siempre fijo
    - Sin placeholders en puntos 1 y 7
    - Remark baterías auto para UN3556/3171/3480/3481
    """)
    st.divider()
    st.markdown("*Desarrollado por Juan Deo · MSC*")

# ── Helper functions ──────────────────────────────────────────────────────────
ISO_MAP = {45.1:"40HC",45.0:"40HC",42.1:"40GP",42.0:"40GP",
           22.1:"20GP",22.0:"20GP",25.1:"20HC",25.0:"20HC",
           86.1:"45HC",86.0:"45HC"}
BATTERY_UNS = {"UN3556","UN3171","UN3480","UN3481"}

def iso_label(code):
    try: return ISO_MAP.get(round(float(code),1), str(code))
    except: return str(code)

def safe(val, fb="N/A"):
    v = str(val).strip() if val is not None else ""
    return fb if v in ("nan","None","NaT","","-") else v

def safe_e(val):
    v = str(val).strip() if val is not None else ""
    return "" if v in ("nan","None","NaT","","-") else v

def imo_fmt(val):
    v = str(val).strip()
    if v in ("nan","None",""): return "N/A"
    try:
        f=float(v); return str(int(f)) if f==int(f) else str(f)
    except: return v

def pg_fmt(val):
    v=str(val).strip(); return "-" if v in ("nan","None","","-") else v

def mp_fmt(val):
    return "P" if str(val).strip() == "P" else "N"

def subsidiary(val):
    v=str(val).strip(); return "N/A" if v in ("-","nan","None","") else v

def flash_pt(row):
    imo=str(row.get("IMOClass","")).strip()
    sub=str(row.get("SubsidiaryRisk","")).strip()
    fp =str(row.get("Flashpoint","")).strip()
    if imo.startswith("3") or "3" in sub.split(","):
        if fp and fp not in ("-","nan","None",""): return f"{fp} C"
    return "N/A"

def state_agg(row):
    imo=str(row.get("IMOClass","")).strip()
    un =str(row.get("UNNumber","")).strip().upper()
    sub=str(row.get("PSNameEnglish","")).upper()
    if un in BATTERY_UNS or "VEHICLE" in sub: return "N/A"
    if imo.startswith("3"): return "Liquid"
    if imo=="5.1":           return "Solid"
    if imo.startswith("5"): return "Liquid"
    if imo.startswith("6"): return "Liquid"
    if imo.startswith("8"): return "Liquid"
    return "N/A"

def outer_pkg(row):
    qty=safe_e(row.get("Qty","")); desc=safe(row.get("PackDescription",""),"N/A")
    if not qty: return desc
    try: return f"{int(float(qty))} x {desc}"
    except: return f"{qty} x {desc}"

def weight_line(row):
    gw=safe_e(row.get("CargoGrossWeight",""))
    nw=safe_e(row.get("IMOWeight",""))
    try: nv=f"{int(float(nw))} kg" if nw else "N/A"
    except: nv=nw or "N/A"
    try:
        if gw:   gv=f"{int(float(gw))} kg"
        elif nw: gv=f"{int(float(nw)*1.05)} kg (approx.)"
        else:    gv="N/A"
    except: gv=gw or "N/A"
    return f"Gross Weight: {gv} / Net Weight: {nv}"

def ts_ports(row):
    pod=safe_e(row.get("POD_Final",""))
    parts=[safe_e(row.get(f,"")) for f in
           ["TranshipmentPort1","TranshipmentPort2","TranshipmentPort3"]]
    return " / ".join(p for p in parts if p and p!=pod) or "N/A"

def emergency_line(row):
    ph=safe(row.get("EmergencyPhone",""),"")
    cn=safe(row.get("Contactname",""),"")
    pr=safe(row.get("EmergencyProvider",""),"")
    parts=[p for p in [ph,cn,pr] if p]
    return " / ".join(parts) if parts else "N/A"

def consolidate_ctrs(rows_df):
    seen=set(); groups={}
    for _,r in rows_df.iterrows():
        lbl=iso_label(r["CtrISOCode"]); ctr=str(r["CtrNumber"]).strip()
        if ctr not in seen:
            seen.add(ctr); groups.setdefault(lbl,[]).append(ctr)
    return " / ".join(f"{l} - {', '.join(c)}" for l,c in groups.items()) or "N/A"

def generate_approval(booking_ref, rows_df, vessel, pol):
    r=rows_df.iloc[0]; un=safe(r.get("UNNumber",""),"").upper()
    lines=["Good day.",""]
    if un in BATTERY_UNS: lines+=["Please note the batteries are new.",""]
    lines.append("1) Reference Number / CSO Number:")
    lines.append(f"2) Name and Voyage Number of Vessel: {vessel}")
    lines.append(f"3) Booking Reference: {booking_ref}")
    pol_val=safe(r.get("POL_First",""),"N/A")
    lines.append(f"4) Place of Receipt: {pol_val}")
    lines.append(f"5) Port of Loading: {pol}")
    lines.append(f"6) Port of Transshipment (if necessary): {ts_ports(r)}")
    lines.append("7) Name & Voyage of Transshipment (if necessary):")
    pod=safe(r.get("POD_Final",""),"N/A")
    lines.append(f"8) Port of Discharge: {pod}")
    lines.append(f"9) Port of Delivery: {safe(r.get('DischargeTerminal',''),pod)}")
    lines.append(f"10) Type, Size and Number of Containers: {consolidate_ctrs(rows_df)}")
    lines.append(f"11) Proper Shipping Name (as named in IMDG code): {safe(r.get('SubstNameEnglish',''),'N/A').upper()}")
    lines.append(f"12) Commodity (Correct Technical Name): {safe(r.get('PSNameEnglish',''),'N/A')}")
    lines.append(f"13) State of Aggregate (Solid, Liquid, Gas, Viscous): {state_agg(r)}")
    imo=imo_fmt(r.get("IMOClass","")); pg=pg_fmt(r.get("PG",""))
    lines.append(f"14) IMO Class / UN No / IMDG Packaging Group: {imo} / {un} / {pg}")
    inq=safe_e(r.get("InnerQty","")); ind=safe_e(r.get("InnerPackingDescription",""))
    inns=""
    try:
        if inq and ind: inns=f"{int(float(inq))} x {ind}"
    except: pass
    lines.append(f"15) Number & Type of Inner Package: {inns}")
    lines.append(f"16) Number & Type of Outer Package / Package Code: {outer_pkg(r)}")
    lines.append(f"17) Quantity (Gross Weight / Net Weight): {weight_line(r)}")
    lines.append("18) Net explosive quantity (NEQ) / Net explosive contents (NEC): N/A")
    lines.append(f"19) Subsidiary Risk(s): {subsidiary(r.get('SubsidiaryRisk',''))}")
    lines.append(f"20) Flash Point (for IMDG Class 3 or Subsidiary Risk 3): {flash_pt(r)}")
    sadt=safe_e(r.get("CAA",""))
    lines.append(f"21) SADT (Self-accelerating Decomposition Temperature): {sadt or 'N/A'}")
    lines.append(f"22) Marine Pollutant (N=no, P=Marine Pollutant): {mp_fmt(r.get('MP',''))}")
    lines.append(f"23) Emergency Schedule code (EMS CODE): {safe(r.get('EmSCode',''),'N/A')}")
    lines.append("24) Limited Quantity (Yes or No): No")
    lines.append("25) Excepted quantities (Yes or No): No")
    lines.append(f"26) Emergency Response Telephone Number (as required by regulations) + Company + Contact Name: {emergency_line(r)}")
    seg=safe_e(r.get("SegregationGroup",""))
    lines.append(f"27) Group: {seg or 'N/A'}")
    lines.append("28) Other Special Requirement: N/A")
    return "\n".join(lines)

def build_excel(df, vessel, pol):
    bookings=list(dict.fromkeys(df["BookingRef"].tolist()))
    DARK="1B2A4A"; MID="2E5090"; LT="D6E4F7"; ACC="F0A500"
    WHITE="FFFFFF"; LGREY="F5F7FA"
    thin=Side(style="thin",color="BDD1EE"); med=Side(style="medium",color=MID)
    def mfill(c): return PatternFill("solid",fgColor=c)
    def bdr(): return Border(left=thin,right=thin,top=thin,bottom=thin)

    wb=openpyxl.Workbook()

    # DATA_SOURCE
    ws=wb.active; ws.title="DATA_SOURCE"
    ws.merge_cells("A1:AY1"); c=ws["A1"]
    c.value=f"  DATA_SOURCE  |  Vessel: {vessel}  |  POL: {pol}"
    c.font=Font(name="Arial",size=11,bold=True,color=WHITE)
    c.fill=mfill(DARK); c.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    ws.row_dimensions[1].height=22
    cols=list(df.columns)
    for ci,col in enumerate(cols,1):
        cell=ws.cell(row=2,column=ci,value=col)
        cell.font=Font(name="Arial",size=9,bold=True,color=WHITE); cell.fill=mfill(MID)
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        cell.border=Border(left=med,right=med,top=med,bottom=med)
    ws.row_dimensions[2].height=34
    for ri,(_,row) in enumerate(df.iterrows(),3):
        rf=mfill(LGREY) if ri%2==0 else mfill(WHITE)
        for ci,col in enumerate(cols,1):
            val=row[col]
            try:
                if hasattr(val,"strftime"): val=val.strftime("%Y-%m-%d")
                elif str(val) in ("nan","NaT"): val=""
            except: val=""
            cell=ws.cell(row=ri,column=ci,value=val)
            cell.font=Font(name="Arial",size=9); cell.fill=rf
            cell.alignment=Alignment(vertical="center"); cell.border=bdr()
        ws.row_dimensions[ri].height=15
    for ci,col in enumerate(cols,1):
        mx=max(len(str(col)),*(len(str(ws.cell(row=ri,column=ci).value or "")) for ri in range(3,ws.max_row+1)))
        ws.column_dimensions[get_column_letter(ci)].width=min(mx+2,28)
    ws.freeze_panes="A3"

    # OUTPUT_TEMPLATE
    wo=wb.create_sheet("OUTPUT_TEMPLATE")
    wo.column_dimensions["A"].width=26; wo.column_dimensions["B"].width=100
    wo.merge_cells("A1:B1"); c=wo["A1"]
    c.value=f"  DG APPROVAL TEMPLATE  |  {datetime.now().strftime('%d %b %Y %H:%M')}  |  {len(bookings)} booking(s)  |  {vessel}"
    c.font=Font(name="Arial",size=11,bold=True,color=WHITE); c.fill=mfill(DARK)
    c.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    wo.row_dimensions[1].height=22; outR=2

    for bki,booking in enumerate(bookings):
        rows_df=df[df["BookingRef"]==booking].copy()
        r0=rows_df.iloc[0]
        unique_ctrs=list(dict.fromkeys(rows_df["CtrNumber"].tolist()))
        unique_uns=list(dict.fromkeys(rows_df["UNNumber"].tolist()))

        wo.merge_cells(f"A{outR}:B{outR}")
        hdr=wo[f"A{outR}"]
        hdr.value=f"  BOOKING #{bki+1}  |  {booking}  |  {len(unique_ctrs)} container(s)  |  {len(unique_uns)} UN class(es)"
        hdr.font=Font(name="Arial",size=10,bold=True,color=WHITE); hdr.fill=mfill(MID)
        hdr.alignment=Alignment(horizontal="left",vertical="center",indent=1)
        wo.row_dimensions[outR].height=18; outR+=1

        for uni,un_val in enumerate(unique_uns):
            un_rows=rows_df[rows_df["UNNumber"]==un_val] if len(unique_uns)>1 else rows_df
            if len(unique_uns)>1:
                wo.merge_cells(f"A{outR}:B{outR}")
                uh=wo[f"A{outR}"]
                psn=safe(un_rows.iloc[0].get("SubstNameEnglish",""),"N/A").upper()
                uh.value=f"    UN Class {uni+1} of {len(unique_uns)}  |  {un_val}  -  {psn}"
                uh.font=Font(name="Arial",size=9,bold=True,color=WHITE); uh.fill=mfill(ACC)
                uh.alignment=Alignment(horizontal="left",vertical="center",indent=2)
                wo.row_dimensions[outR].height=16; outR+=1

            text=generate_approval(booking, un_rows if len(unique_uns)>1 else rows_df, vessel, pol)
            mp=mp_fmt(un_rows.iloc[0].get("MP",""))

            lc=wo[f"A{outR}"]
            lc.value="\n".join([
                f"Vessel:  {vessel}",
                f"UN No:   {un_val}",
                f"Class:   {imo_fmt(un_rows.iloc[0].get('IMOClass',''))}",
                f"PSN:     {safe(un_rows.iloc[0].get('SubstNameEnglish',''),'N/A')[:28]}",
                f"CTRs:    {len(unique_ctrs)}",
                f"MP:      {mp}",
                f"POD:     {safe(r0.get('POD_Final',''),'N/A')}",
            ])
            lc.font=Font(name="Courier New",size=9,color=DARK)
            lc.fill=mfill(LT); lc.alignment=Alignment(vertical="top",wrap_text=True,indent=1)

            tc=wo[f"B{outR}"]
            tc.value=text
            tc.font=Font(name="Courier New",size=9.5,color="1E1E1E")
            tc.fill=mfill(WHITE); tc.alignment=Alignment(vertical="top",wrap_text=True,indent=1)
            wo.row_dimensions[outR].height=max(text.count("\n")*14,90)
            outR+=2

    wo.merge_cells(f"A{outR}:B{outR}"); ft=wo[f"A{outR}"]
    ft.value=f"  {len(bookings)} booking(s)  |  {vessel}  |  Port of Loading: {pol}"
    ft.font=Font(name="Arial",size=9,italic=True,color=WHITE); ft.fill=mfill(DARK)
    ft.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    wo.row_dimensions[outR].height=16; wo.freeze_panes="A2"

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf, bookings

# ── Main UI ─────────────────────────────────────────────────────────────
col1, col2 = st.columns([1, 1], gap="large")

with col1:
    st.markdown("### 📂 1. Sube el DCR")
    uploaded = st.file_uploader(
        "Arrastra tu archivo DCR aquí",
        type=["xls","xlsx"],
        help="Formato DMAXS — sheet 'BR', headers en fila 3"
    )

with col2:
    st.markdown("### 🔧 2. Revisa la configuración")
    st.markdown(f"**Barco:** `{vessel_name}`")
    st.markdown(f"**Port of Loading:** `{port_loading}`")
    st.caption("Edita estos valores en el panel izquierdo si cambiaron.")

st.divider()

if uploaded:
    # Load data directly without soffice conversion
    try:
        # pandas can read .xls files directly with xlrd
        df = pd.read_excel(uploaded, sheet_name="BR", header=2)
        df = df.dropna(subset=["BookingRef"])
    except Exception as e:
        st.error(f"❌ Error leyendo el DCR: {e}")
        st.info("📋 Asegúrate que:\n- El archivo sea .xls o .xlsx válido\n- Exista la sheet 'BR'\n- Los headers estén en fila 3")
        st.stop()

    bookings      = list(dict.fromkeys(df["BookingRef"].tolist()))
    unique_uns    = df["UNNumber"].nunique()
    battery_books = [b for b in bookings
                     if df[df["BookingRef"]==b]["UNNumber"].iloc[0].upper() in BATTERY_UNS]
    mp_p_count    = sum(1 for _,r in df.drop_duplicates("BookingRef").iterrows()
                        if mp_fmt(r.get("MP",""))=="P")

    # Preview metrics
    st.markdown("### 📊 3. Preview del DCR")
    m1,m2,m3,m4 = st.columns(4)
    m1.metric("Bookings",    len(bookings))
    m2.metric("Containers",  df["CtrNumber"].nunique())
    m3.metric("Marine Poll.", f"{mp_p_count} P / {len(bookings)-mp_p_count} N")
    m4.metric("Con baterías", len(battery_books))

    # Booking table preview
    with st.expander("Ver detalle de bookings", expanded=False):
        preview_rows=[]
        for bk in bookings:
            rows=df[df["BookingRef"]==bk]
            r0=rows.iloc[0]
            preview_rows.append({
                "Booking":    bk,
                "UN(s)":     ", ".join(dict.fromkeys(rows["UNNumber"].tolist())),
                "CTRs":      ", ".join(dict.fromkeys(rows["CtrNumber"].tolist()))[:60],
                "MP":        mp_fmt(r0.get("MP","")),
                "Baterías":  "✅" if r0.get("UNNumber","").upper() in BATTERY_UNS else "—",
                "POD":       safe(r0.get("POD_Final",""),"N/A"),
            })
        st.dataframe(pd.DataFrame(preview_rows), use_container_width=True, hide_index=True)

    st.divider()
    st.markdown("### 🚀 4. Generar plantillas")

    if st.button("⚡ GENERATE APPROVALS", use_container_width=True):
        with st.spinner(f"Generando {len(bookings)} plantillas..."):
            output_buf, bk_list = build_excel(df, vessel_name, port_loading)

        st.markdown(f"""
        <div class="success-box">
        ✅ <strong>{len(bk_list)} plantillas generadas correctamente.</strong><br>
        Vessel: <code>{vessel_name}</code> · Port of Loading: <code>{port_loading}</code>
        </div>
        """, unsafe_allow_html=True)

        dcr_name = os.path.splitext(uploaded.name)[0]
        st.download_button(
            label="📥 Descargar DG_Approval_Generator.xlsx",
            data=output_buf,
            file_name=f"DG_Approval_{dcr_name}_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

else:
    st.markdown("""
    <div class="warning-box">
    📌 <strong>Sube un archivo DCR</strong> para comenzar.<br>
    Formato esperado: exportación DMAXS (.xls o .xlsx), sheet <code>BR</code>, headers en fila 3.
    </div>
    """, unsafe_allow_html=True)

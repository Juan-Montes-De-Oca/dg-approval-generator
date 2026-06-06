"""
DG APPROVAL GENERATOR
=====================
Uso: python generate_approvals.py  ruta/al/DCR.xls

El script convierte el DCR a .xlsx si es necesario,
genera todas las plantillas y guarda el resultado en
DG_Approval_Generator.xlsx en la misma carpeta.

Configuracion:
  Edita las variables de la seccion CONFIG mas abajo.
"""

import pandas as pd, openpyxl, subprocess, os, sys
from openpyxl.styles import Font, PatternFill, Alignment, Side, Border
from openpyxl.utils import get_column_letter
from datetime import datetime

# ================================================================
# CONFIG — edita estos valores cuando cambie el barco
# ================================================================
VESSEL_NAME  = "AS SVENJA / V.26022N"   # nombre del barco y voyage
PORT_LOADING = "BALBOA"                  # puerto de carga (fijo)
# ================================================================

ISO_MAP = {45.1:"40HC",45.0:"40HC",42.1:"40GP",42.0:"40GP",
           22.1:"20GP",22.0:"20GP",25.1:"20HC",25.0:"20HC",
           86.1:"45HC",86.0:"45HC"}
BATTERY_UNS = {"UN3556","UN3171","UN3480","UN3481"}

def iso_label(code):
    try: return ISO_MAP.get(round(float(code),1), str(code))
    except: return str(code)

def safe(val, fb="N/A"):
    v = str(val).strip() if val is not None else ""
    return fb if v in ("nan","None","NaT","","-") else v

def safe_e(val):
    v = str(val).strip() if val is not None else ""
    return "" if v in ("nan","None","NaT","","-") else v

def imo_fmt(val):
    v = str(val).strip()
    if v in ("nan","None",""): return "N/A"
    try:
        f=float(v); return str(int(f)) if f==int(f) else str(f)
    except: return v

def pg_fmt(val):
    v=str(val).strip(); return "-" if v in ("nan","None","","-") else v

def mp_fmt(val):
    # * = NO es marine pollutant (N), P = SI es marine pollutant (P)
    return "P" if str(val).strip() == "P" else "N"

def subsidiary(val):
    v=str(val).strip(); return "N/A" if v in ("-","nan","None","") else v

def flash_pt(row):
    imo=str(row.get("IMOClass","")).strip()
    sub=str(row.get("SubsidiaryRisk","")).strip()
    fp =str(row.get("Flashpoint","")).strip()
    if imo.startswith("3") or "3" in sub.split(","):
        if fp and fp not in ("-","nan","None",""): return f"{fp} C"
    return "N/A"

def state_agg(row):
    imo=str(row.get("IMOClass","")).strip()
    un =str(row.get("UNNumber","")).strip().upper()
    sub=str(row.get("PSNameEnglish","")).upper()
    if un in BATTERY_UNS or "VEHICLE" in sub: return "N/A"
    if imo.startswith("3"): return "Liquid"
    if imo=="5.1":          return "Solid"
    if imo.startswith("5"): return "Liquid"
    if imo.startswith("6"): return "Liquid"
    if imo.startswith("8"): return "Liquid"
    return "N/A"

def outer_pkg(row):
    qty =safe_e(row.get("Qty",""))
    desc=safe(row.get("PackDescription",""),"N/A")
    if not qty: return desc
    try: return f"{int(float(qty))} x {desc}"
    except: return f"{qty} x {desc}"

def weight_line(row):
    gw=safe_e(row.get("CargoGrossWeight",""))
    nw=safe_e(row.get("IMOWeight",""))
    try: nv=f"{int(float(nw))} kg" if nw else "N/A"
    except: nv=nw or "N/A"
    try:
        if gw:  gv=f"{int(float(gw))} kg"
        elif nw: gv=f"{int(float(nw)*1.05)} kg (approx.)"
        else:    gv="N/A"
    except: gv=gw or "N/A"
    return f"Gross Weight: {gv} / Net Weight: {nv}"

def ts_ports(row):
    pod=safe_e(row.get("POD_Final",""))
    parts=[safe_e(row.get(f,"")) for f in
           ["TranshipmentPort1","TranshipmentPort2","TranshipmentPort3"]]
    return " / ".join(p for p in parts if p and p!=pod) or "N/A"

def emergency_line(row):
    ph=safe(row.get("EmergencyPhone",""),"")
    cn=safe(row.get("Contactname",""),"")
    pr=safe(row.get("EmergencyProvider",""),"")
    parts=[p for p in [ph,cn,pr] if p]
    return " / ".join(parts) if parts else "N/A"

def consolidate_ctrs(rows_df):
    seen=set(); groups={}
    for _,r in rows_df.iterrows():
        lbl=iso_label(r["CtrISOCode"]); ctr=str(r["CtrNumber"]).strip()
        if ctr not in seen:
            seen.add(ctr); groups.setdefault(lbl,[]).append(ctr)
    return " / ".join(f"{l} - {', '.join(c)}" for l,c in groups.items()) or "N/A"

def generate_approval(booking_ref, rows_df):
    r  = rows_df.iloc[0]
    un = safe(r.get("UNNumber",""),"").upper()

    lines = ["Good day.",""]
    if un in BATTERY_UNS:
        lines += ["Please note the batteries are new.",""]

    lines.append("1) Reference Number / CSO Number:")
    lines.append(f"2) Name and Voyage Number of Vessel: {VESSEL_NAME}")
    lines.append(f"3) Booking Reference: {booking_ref}")
    pol = safe(r.get("POL_First",""),"N/A")
    lines.append(f"4) Place of Receipt: {pol}")
    lines.append(f"5) Port of Loading: {PORT_LOADING}")
    lines.append(f"6) Port of Transshipment (if necessary): {ts_ports(r)}")
    lines.append("7) Name & Voyage of Transshipment (if necessary):")
    pod = safe(r.get("POD_Final",""),"N/A")
    lines.append(f"8) Port of Discharge: {pod}")
    lines.append(f"9) Port of Delivery: {safe(r.get('DischargeTerminal',''),pod)}")
    lines.append(f"10) Type, Size and Number of Containers: {consolidate_ctrs(rows_df)}")
    lines.append(f"11) Proper Shipping Name (as named in IMDG code): {safe(r.get('SubstNameEnglish',''),'N/A').upper()}")
    lines.append(f"12) Commodity (Correct Technical Name): {safe(r.get('PSNameEnglish',''),'N/A')}")
    lines.append(f"13) State of Aggregate (Solid, Liquid, Gas, Viscous): {state_agg(r)}")
    imo=imo_fmt(r.get("IMOClass","")); pg=pg_fmt(r.get("PG",""))
    lines.append(f"14) IMO Class / UN No / IMDG Packaging Group: {imo} / {un} / {pg}")
    inq=safe_e(r.get("InnerQty","")); ind=safe_e(r.get("InnerPackingDescription",""))
    inns=""
    try:
        if inq and ind: inns=f"{int(float(inq))} x {ind}"
    except: pass
    lines.append(f"15) Number & Type of Inner Package: {inns}")
    lines.append(f"16) Number & Type of Outer Package / Package Code: {outer_pkg(r)}")
    lines.append(f"17) Quantity (Gross Weight / Net Weight): {weight_line(r)}")
    lines.append("18) Net explosive quantity (NEQ) / Net explosive contents (NEC): N/A")
    lines.append(f"19) Subsidiary Risk(s): {subsidiary(r.get('SubsidiaryRisk',''))}")
    lines.append(f"20) Flash Point (for IMDG Class 3 or Subsidiary Risk 3): {flash_pt(r)}")
    sadt=safe_e(r.get("CAA",""))
    lines.append(f"21) SADT (Self-accelerating Decomposition Temperature): {sadt or 'N/A'}")
    lines.append(f"22) Marine Pollutant (N=no, P=Marine Pollutant): {mp_fmt(r.get('MP',''))}")
    lines.append(f"23) Emergency Schedule code (EMS CODE): {safe(r.get('EmSCode',''),'N/A')}")
    lines.append("24) Limited Quantity (Yes or No): No")
    lines.append("25) Excepted quantities (Yes or No): No")
    lines.append(f"26) Emergency Response Telephone Number (as required by regulations) + Company + Contact Name: {emergency_line(r)}")
    seg=safe_e(r.get("SegregationGroup",""))
    lines.append(f"27) Group: {seg or 'N/A'}")
    lines.append("28) Other Special Requirement: N/A")
    return "\n".join(lines)

def build_excel(source_xlsx, output_path):
    df = pd.read_excel(source_xlsx, sheet_name="BR", header=2)
    df = df.dropna(subset=["BookingRef"])
    bookings = list(dict.fromkeys(df["BookingRef"].tolist()))

    DARK="1B2A4A"; MID="2E5090"; LT="D6E4F7"; ACC="F0A500"
    WHITE="FFFFFF"; LGREY="F5F7FA"
    thin=Side(style="thin",color="BDD1EE"); med=Side(style="medium",color=MID)
    def mfill(c): return PatternFill("solid",fgColor=c)
    def bdr(): return Border(left=thin,right=thin,top=thin,bottom=thin)

    wb = openpyxl.Workbook()

    # DATA_SOURCE
    ws=wb.active; ws.title="DATA_SOURCE"
    ws.merge_cells("A1:AY1"); c=ws["A1"]
    c.value=f"  DATA_SOURCE  |  {source_xlsx}  |  Vessel: {VESSEL_NAME}"
    c.font=Font(name="Arial",size=11,bold=True,color=WHITE)
    c.fill=mfill(DARK); c.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    ws.row_dimensions[1].height=22
    cols=list(df.columns)
    for ci,col in enumerate(cols,1):
        cell=ws.cell(row=2,column=ci,value=col)
        cell.font=Font(name="Arial",size=9,bold=True,color=WHITE)
        cell.fill=mfill(MID)
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        cell.border=Border(left=med,right=med,top=med,bottom=med)
    ws.row_dimensions[2].height=34
    for ri,(_,row) in enumerate(df.iterrows(),3):
        rf=mfill(LGREY) if ri%2==0 else mfill(WHITE)
        for ci,col in enumerate(cols,1):
            val=row[col]
            try:
                if hasattr(val,"strftime"): val=val.strftime("%Y-%m-%d")
                elif str(val) in ("nan","NaT"): val=""
            except: val=""
            cell=ws.cell(row=ri,column=ci,value=val)
            cell.font=Font(name="Arial",size=9); cell.fill=rf
            cell.alignment=Alignment(vertical="center"); cell.border=bdr()
        ws.row_dimensions[ri].height=15
    for ci,col in enumerate(cols,1):
        mx=max(len(str(col)),*(len(str(ws.cell(row=ri,column=ci).value or "")) for ri in range(3,ws.max_row+1)))
        ws.column_dimensions[get_column_letter(ci)].width=min(mx+2,28)
    ws.freeze_panes="A3"

    # OUTPUT_TEMPLATE
    wo=wb.create_sheet("OUTPUT_TEMPLATE")
    wo.column_dimensions["A"].width=26; wo.column_dimensions["B"].width=100
    wo.merge_cells("A1:B1"); c=wo["A1"]
    c.value=f"  DG APPROVAL TEMPLATE  |  {datetime.now().strftime('%d %b %Y')}  |  {len(bookings)} booking(s)  |  {VESSEL_NAME}"
    c.font=Font(name="Arial",size=11,bold=True,color=WHITE); c.fill=mfill(DARK)
    c.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    wo.row_dimensions[1].height=22
    outR=2

    for bki,booking in enumerate(bookings):
        rows_df = df[df["BookingRef"]==booking].copy()
        r0      = rows_df.iloc[0]
        unique_ctrs = list(dict.fromkeys(rows_df["CtrNumber"].tolist()))
        unique_uns  = list(dict.fromkeys(rows_df["UNNumber"].tolist()))

        wo.merge_cells(f"A{outR}:B{outR}")
        hdr=wo[f"A{outR}"]
        hdr.value=f"  BOOKING #{bki+1}  |  {booking}  |  {len(unique_ctrs)} container(s)  |  {len(unique_uns)} UN class(es)"
        hdr.font=Font(name="Arial",size=10,bold=True,color=WHITE); hdr.fill=mfill(MID)
        hdr.alignment=Alignment(horizontal="left",vertical="center",indent=1)
        wo.row_dimensions[outR].height=18; outR+=1

        for uni,un_val in enumerate(unique_uns):
            un_rows = rows_df[rows_df["UNNumber"]==un_val] if len(unique_uns)>1 else rows_df

            if len(unique_uns)>1:
                wo.merge_cells(f"A{outR}:B{outR}")
                uh=wo[f"A{outR}"]
                psn=safe(un_rows.iloc[0].get("SubstNameEnglish",""),"N/A").upper()
                uh.value=f"    UN Class {uni+1} of {len(unique_uns)}  |  {un_val}  -  {psn}"
                uh.font=Font(name="Arial",size=9,bold=True,color=WHITE); uh.fill=mfill(ACC)
                uh.alignment=Alignment(horizontal="left",vertical="center",indent=2)
                wo.row_dimensions[outR].height=16; outR+=1

            text = generate_approval(booking, un_rows if len(unique_uns)>1 else rows_df)
            mp   = mp_fmt(un_rows.iloc[0].get("MP",""))

            lc=wo[f"A{outR}"]
            lc.value="\n".join([
                f"Vessel:  {VESSEL_NAME}",
                f"UN No:   {un_val}",
                f"Class:   {imo_fmt(un_rows.iloc[0].get('IMOClass',''))}",
                f"PSN:     {safe(un_rows.iloc[0].get('SubstNameEnglish',''),'N/A')[:28]}",
                f"CTRs:    {len(unique_ctrs)}",
                f"MP:      {mp}",
                f"POD:     {safe(r0.get('POD_Final',''),'N/A')}",
            ])
            lc.font=Font(name="Courier New",size=9,color=DARK)
            lc.fill=mfill(LT); lc.alignment=Alignment(vertical="top",wrap_text=True,indent=1)

            tc=wo[f"B{outR}"]
            tc.value=text
            tc.font=Font(name="Courier New",size=9.5,color="1E1E1E")
            tc.fill=mfill(WHITE); tc.alignment=Alignment(vertical="top",wrap_text=True,indent=1)
            wo.row_dimensions[outR].height=max(text.count("\n")*14,90)
            outR+=2

    wo.merge_cells(f"A{outR}:B{outR}"); ft=wo[f"A{outR}"]
    ft.value=f"  {len(bookings)} booking(s)  |  {VESSEL_NAME}  |  Port of Loading: {PORT_LOADING}"
    ft.font=Font(name="Arial",size=9,italic=True,color=WHITE); ft.fill=mfill(DARK)
    ft.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    wo.row_dimensions[outR].height=16; wo.freeze_panes="A2"

    wb.save(output_path)
    print(f"\nSaved: {output_path}")
    print(f"Bookings: {len(bookings)}")
    for bk in bookings:
        rows=df[df["BookingRef"]==bk]
        ctrs=list(dict.fromkeys(rows["CtrNumber"].tolist()))
        uns =list(dict.fromkeys(rows["UNNumber"].tolist()))
        mp  =mp_fmt(rows.iloc[0].get("MP",""))
        print(f"  {bk} | MP={mp} | CTRs={len(ctrs)} | UNs={uns}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python generate_approvals.py ruta/al/DCR.xls")
        sys.exit(1)

    src = sys.argv[1]
    if not os.path.exists(src):
        print(f"Error: no se encontro el archivo {src}"); sys.exit(1)

    # Convertir .xls a .xlsx si es necesario
    if src.lower().endswith(".xls"):
        print(f"Convirtiendo {os.path.basename(src)} a xlsx...")
        out_dir = os.path.dirname(os.path.abspath(src))
        subprocess.run(["soffice","--headless","--convert-to","xlsx",
                        src,"--outdir",out_dir], capture_output=True)
        src = os.path.splitext(src)[0]+".xlsx"
        if not os.path.exists(src):
            print("Error: fallo la conversion. Instala LibreOffice."); sys.exit(1)
        print(f"Convertido: {src}")

    out = os.path.join(os.path.dirname(os.path.abspath(sys.argv[1])),
                       "DG_Approval_Generator.xlsx")
    build_excel(src, out)
